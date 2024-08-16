import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook("testdata.xlsx")
sheet = wb.active

# Set up the Selenium WebDriver (make sure you have the appropriate driver installed)
driver = webdriver.Chrome()  # You can use a different driver if needed

# Iterate through rows in the Excel file
for row in sheet.iter_rows(
    min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
):
    part_name = row[3].value + " digikey"
    if part_name:
        # Perform Google search
        driver.get("https://www.google.com")
        search_box = driver.find_element(By.NAME, "q")
        search_box.send_keys(part_name)
        search_box.send_keys(Keys.RETURN)

        # Wait for search results to load
        wait = WebDriverWait(driver, 10)
        first_result = wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "div.g > div > div > div > a")
            )
        )

        # Get the first link
        first_link = first_result.get_attribute("href")

        # Save the link to the Excel file
        row[1].value = first_link

    # Add a small delay to avoid overwhelming Google with requests
    time.sleep(2)

# Save the updated Excel file
wb.save("parts_updated.xlsx")

# Close the browser
# driver.quit()

print("Process completed. Results saved in 'parts_updated.xlsx'")
# using selenium, access an excel file and google search the part names in it and save the first link to the excel
